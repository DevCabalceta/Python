from openpyxl import load_workbook, Workbook

nombre_Archivo = r"C:\Users\Gabriel Cabalceta\Documents\Python\CLASE-11\EjemploInformacion.xlsx"
nombre_hoja = "EjercicioNuevo"

wb = load_workbook(nombre_Archivo)

#Si la hoja no existe la creamos
if nombre_hoja not in wb.sheetnames:
    ws = wb.create_sheet(nombre_hoja)
else:
    ws = wb[nombre_hoja]

ws["B5"] = "VLA Hola"

wb.save(nombre_Archivo)