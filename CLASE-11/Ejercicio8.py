from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

class Persona:
    def __init__(self,nombre, apellido, edad):
        self.nombre = nombre
        self.apellido = apellido
        self.edad = edad


nombre_Archivo = r"C:\Users\Gabriel Cabalceta\Documents\Python\CLASE-11\EjemploInformacion.xlsx"
hojaLecturaPersonas = "Personas"

wb = load_workbook(nombre_Archivo)
ws = wb[hojaLecturaPersonas]

vectorPersonas = []

for fila in ws.iter_rows(min_row=2, min_col=1, values_only=True):
    Nombre, Apellido, Edad = fila
    persona = Persona(Nombre, Apellido, Edad)
    vectorPersonas.append(persona)


# Jugar con los datos de vector 
# Van a crear una hoja de estadisticas 
# En la B2 van a poner la cantidad de personas del vector 
hoja_nueva = "Estadisticas"
cantidadPersonas = (len(vectorPersonas))

ws = load_workbook(nombre_Archivo)

#Si la hoja no existe la creamos
if hoja_nueva not in wb.sheetnames:
    ws = wb.create_sheet(hoja_nueva)
else:
    ws = wb[hoja_nueva]

ws.merge_cells("B1:D1")
ws["B1"] = "Cantidad de Personas"
ws["B2"] = cantidadPersonas

# Ponerle color de la b1 a la d1, negrita y centrar
ws["B1"].font = Font(bold=True, color="FFFFFF")
ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
ws["B1"].fill = PatternFill(
    start_color="00B050",
    end_color="00B050",
    fill_type="solid"
)

# Escribir n registros a partir de una celda, desde la fila 10
filaInicio = 10
ws[f"A{filaInicio}"] = "Nombre"
ws[f"B{filaInicio}"] = "Apellido"
ws[f"C{filaInicio}"] = "Edad"

# Poner todas esas palabras en negrita
for col in["A", "B", "C"]:
    ws[f"{col}{filaInicio}"].font = Font(bold=True, color="FFFFFF")
    ws[f"{col}{filaInicio}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{col}{filaInicio}"].fill = PatternFill(
        start_color="00B050",
        end_color="00B050",
        fill_type="solid"
    )

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 50

filaActual = filaInicio + 1
for persona in vectorPersonas:
    ws[f"A{filaActual}"] = persona.nombre
    ws[f"B{filaActual}"] = persona.apellido
    ws[f"C{filaActual}"] = persona.edad
    filaActual = filaActual + 1
    
wb.save(nombre_Archivo)



